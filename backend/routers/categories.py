"""Product categories — dynamic category master for the Product module.

A product-facing category catalog (distinct from the accounting `stock-categories`
master): the New Product form picks from these, and each product stores both the
category name (back-compat / display) and a `category_id` FK.

Soft-delete only (status flips to Inactive + is_deleted flag); a category that
still has products linked cannot be deleted. Supports search, sort by
display_order, auto-generated codes, nestable parents, bulk Excel/CSV import and
Excel export.

Mounted at the app root: /categories.
"""
import csv
import io
import re
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from pydantic import BaseModel, Field

from sqlalchemy import select

from core import cache
from core.auth_utils import get_current_user, require_admin
from core.db import db, get_session
from core.models import ProductCategory, ProductCategoryUpdate
from core.schema import Product
from core.utils import crud_create, crud_get, crud_update, new_id, now_iso

router = APIRouter(tags=["categories"])

COLL = "product_categories"

# Common scaffolding categories offered as one-click quick-add templates.
DEFAULT_TEMPLATES = [
    "Cuplock", "Ringlock", "Frame", "Planks", "Accessories", "Couplers",
    "Base Jack", "U Head Jack", "Steel Prop", "Guard Rail", "Other",
]


def _slug_code(name: str) -> str:
    """Derive a short uppercase code from the name (e.g. 'U Head Jack' → 'UHJ').

    Initials of each word when multi-word; otherwise the first letters. Always
    made unique by the caller via _unique_code.
    """
    words = re.findall(r"[A-Za-z0-9]+", name.upper())
    if not words:
        return "CAT"
    if len(words) >= 2:
        base = "".join(w[0] for w in words)[:6]
    else:
        base = words[0][:6]
    return base or "CAT"


async def _unique_code(base: str, exclude_id: Optional[str] = None) -> str:
    """Return `base`, or base+N, guaranteed unique across non-deleted categories."""
    candidate = base
    n = 1
    while True:
        q: dict[str, Any] = {"code": candidate, "is_deleted": {"$ne": True}}
        if exclude_id:
            q["id"] = {"$ne": exclude_id}
        if not await db[COLL].find_one(q, {"_id": 0, "id": 1}):
            return candidate
        n += 1
        candidate = f"{base}{n}"


async def _assert_name_unique(name: str, exclude_id: Optional[str] = None) -> None:
    q: dict[str, Any] = {
        "name": {"$regex": f"^{re.escape(name.strip())}$", "$options": "i"},
        "is_deleted": {"$ne": True},
    }
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    if await db[COLL].find_one(q, {"_id": 0, "id": 1}):
        raise HTTPException(400, f"A category named '{name}' already exists.")


async def _link_count(category_id: str, name: str) -> int:
    """How many products reference this category (by id or by legacy name)."""
    return await db.products.count_documents(
        {"$or": [{"category_id": category_id}, {"category": name}]}
    )


async def _product_counts_by_category(rows: list[dict]) -> dict[str, int]:
    """Batch product link counts — one query instead of N per category."""
    if not rows:
        return {}
    counts = {r["id"]: 0 for r in rows if r.get("id")}
    name_to_id = {r.get("name", ""): r["id"] for r in rows if r.get("name")}
    async with get_session() as session:
        result = await session.execute(select(Product.category_id, Product.category))
        for cat_id, cat_name in result.all():
            targets: set[str] = set()
            if cat_id and cat_id in counts:
                targets.add(cat_id)
            if cat_name and cat_name in name_to_id:
                targets.add(name_to_id[cat_name])
            for tid in targets:
                counts[tid] += 1
    return counts


# ───────────────────────── List / search / sort ─────────────────────────

@router.get("/categories")
async def list_categories(
    q: Optional[str] = None,
    status: Optional[str] = None,
    include_inactive: bool = True,
    sort: str = "display_order",
    _: dict = Depends(get_current_user),
):
    """List categories (excludes soft-deleted). Search by name/code/description,
    optional status filter, sorted by display_order (then name) by default."""
    # Hot path: the unfiltered default list loads on most product/inventory
    # screens and is otherwise 2 cross-region round-trips (rows + product counts).
    # Cache it briefly, keyed by the params that change the result. Searches (q
    # set) are varied and infrequent, so they bypass the cache. Writes invalidate
    # the "categories:" prefix via the compat shim's _invalidate_caches.
    cache_key = None
    if not q:
        cache_key = f"categories:list:{status}:{include_inactive}:{sort}"
        cached = cache.get(cache_key)
        if cached is not cache._MISS:
            return cached

    filt: dict = {"is_deleted": {"$ne": True}}
    if status:
        filt["status"] = status
    elif not include_inactive:
        filt["status"] = "Active"
    if q:
        filt["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"code": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
        ]
    sort_field = sort if sort in ("display_order", "name", "code", "created_at", "updated_at") else "display_order"
    rows = await db[COLL].find(filt, {"_id": 0}).sort(
        [(sort_field, 1), ("name", 1)]
    ).to_list(5000)
    # Annotate with product link counts so the UI can show usage / block deletes.
    link_counts = await _product_counts_by_category(rows)
    for r in rows:
        r["product_count"] = link_counts.get(r["id"], 0)
    if cache_key is not None:
        cache.set(cache_key, rows, cache.TTL_REFERENCE)
    return rows


@router.get("/categories/templates")
async def category_templates(_: dict = Depends(get_current_user)):
    """Quick-add scaffolding templates, flagged with whether they already exist."""
    existing = {
        (c.get("name") or "").strip().lower()
        async for c in db[COLL].find({"is_deleted": {"$ne": True}}, {"_id": 0, "name": 1})
    }
    return [{"name": t, "exists": t.strip().lower() in existing} for t in DEFAULT_TEMPLATES]


@router.get("/categories/export")
async def export_categories(_: dict = Depends(get_current_user)):
    """Download all categories as an .xlsx workbook."""
    from openpyxl import Workbook

    rows = await db[COLL].find({"is_deleted": {"$ne": True}}, {"_id": 0}).sort(
        [("display_order", 1), ("name", 1)]
    ).to_list(5000)
    by_id = {r["id"]: r.get("name", "") for r in rows}

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(500, "Could not create workbook sheet")
    ws.title = "Categories"
    headers = ["Name", "Code", "Description", "Parent Category", "Status",
               "Display Order", "Created At", "Updated At"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("name", ""),
            r.get("code", ""),
            r.get("description", "") or "",
            by_id.get(r.get("parent_id"), "") if r.get("parent_id") else "",
            r.get("status", "Active"),
            r.get("display_order", 0),
            r.get("created_at", ""),
            r.get("updated_at", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="categories.xlsx"'},
    )


@router.get("/categories/{item_id}")
async def get_category(item_id: str, _: dict = Depends(get_current_user)):
    return await crud_get(COLL, item_id, label="Category")


# ───────────────────────── Create / update ─────────────────────────

async def _build_category_doc(payload: ProductCategory, exclude_id: Optional[str] = None) -> dict:
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(400, "Category name is required.")
    await _assert_name_unique(name, exclude_id=exclude_id)
    if payload.parent_id:
        await crud_get(COLL, payload.parent_id, label="Parent category")
        if payload.parent_id == exclude_id:
            raise HTTPException(400, "A category cannot be its own parent.")
    code = (payload.code or "").strip() or await _unique_code(_slug_code(name), exclude_id=exclude_id)
    return {
        "name": name,
        "code": code,
        "description": payload.description,
        "parent_id": payload.parent_id or None,
        "status": payload.status or "Active",
        "display_order": payload.display_order or 0,
    }


@router.post("/categories")
async def create_category(payload: ProductCategory, user: dict = Depends(get_current_user)):
    doc = await _build_category_doc(payload)
    doc["is_deleted"] = False
    return await crud_create(COLL, doc, user=user)


@router.put("/categories/{item_id}")
async def update_category(item_id: str, payload: ProductCategoryUpdate, user: dict = Depends(get_current_user)):
    existing = await crud_get(COLL, item_id, label="Category")
    data = {k: v for k, v in payload.model_dump(exclude_unset=True).items()}

    if "name" in data and data["name"] is not None:
        data["name"] = data["name"].strip()
        await _assert_name_unique(data["name"], exclude_id=item_id)
    if "parent_id" in data and data["parent_id"]:
        if data["parent_id"] == item_id:
            raise HTTPException(400, "A category cannot be its own parent.")
        await crud_get(COLL, data["parent_id"], label="Parent category")
    if "code" in data and data["code"]:
        data["code"] = await _unique_code(data["code"].strip(), exclude_id=item_id)
    if "display_order" in data and data["display_order"] is not None:
        data["display_order"] = int(data["display_order"])

    updated = await crud_update(COLL, item_id, data, user=user)

    # Keep the denormalised category name on linked products in sync on rename.
    if "name" in data and data["name"] and data["name"] != existing.get("name"):
        await db.products.update_many(
            {"category_id": item_id}, {"$set": {"category": data["name"], "updated_at": now_iso()}}
        )
    return updated


async def _delete_one_category(item_id: str, force: bool, user: dict) -> dict:
    """Soft-delete a single category, applying the same guards as the public
    endpoint. Raises HTTPException(400) when blocked (sub-categories, or linked
    products without ``force``); returns the result dict on success. Shared by
    the single DELETE and the bulk-delete endpoint so the rules stay identical.
    """
    cat = await crud_get(COLL, item_id, label="Category")
    if await db[COLL].find_one({"parent_id": item_id, "is_deleted": {"$ne": True}}, {"_id": 0, "id": 1}):
        raise HTTPException(400, "Cannot delete: this category has sub-categories. Reassign or delete them first.")
    linked = await _link_count(item_id, cat.get("name", ""))
    unlinked = 0
    if linked:
        if not force:
            raise HTTPException(
                400,
                f"Cannot delete '{cat.get('name')}': {linked} product(s) are linked to it. "
                "Reassign those products to another category first.",
            )
        # Force: detach every linked product (by id or legacy name) from the category.
        res = await db.products.update_many(
            {"$or": [{"category_id": item_id}, {"category": cat.get("name", "")}]},
            {"$set": {"category_id": None, "category": "", "updated_at": now_iso()}},
        )
        unlinked = res.modified_count
    await crud_update(COLL, item_id, {"is_deleted": True, "status": "Inactive"}, user=user)
    return {"ok": True, "soft_deleted": True, "unlinked_products": unlinked}


class BulkDeleteIn(BaseModel):
    ids: list[str] = Field(..., min_length=1, description="Category ids to delete.")
    force: bool = Field(False, description="Unlink linked products and delete anyway.")


@router.post("/categories/bulk-delete")
async def bulk_delete_categories(payload: BulkDeleteIn, user: dict = Depends(require_admin)):
    """Soft-delete many categories at once.

    Each id is processed with the SAME rules as the single delete (blocked by
    sub-categories, and by linked products unless ``force=true``). One bad id
    does not abort the rest: the response reports per-id success/failure plus a
    summary, so the UI can show exactly which ones were skipped and why.

    Order matters when deleting a parent and its children together: children are
    deleted before parents (so a parent isn't blocked by a child that's also in
    the batch). Duplicate ids are de-duplicated.
    """
    seen: set[str] = set()
    ids = [i for i in payload.ids if i and not (i in seen or seen.add(i))]

    # Children before parents: a parent in the batch shouldn't be blocked by a
    # child that's also being deleted in the same call. Sort so any id whose
    # parent_id is ALSO in the batch comes first.
    parent_of: dict[str, Optional[str]] = {}
    if ids:
        rows = await db[COLL].find(
            {"id": {"$in": ids}}, {"_id": 0, "id": 1, "parent_id": 1}
        ).to_list(len(ids))
        parent_of = {r["id"]: r.get("parent_id") for r in rows}
    batch = set(ids)
    # Stable order: items whose parent is in the batch first (depth-first-ish).
    ids.sort(key=lambda i: 0 if parent_of.get(i) in batch else 1)

    results: list[dict] = []
    deleted = 0
    total_unlinked = 0
    for cid in ids:
        try:
            res = await _delete_one_category(cid, payload.force, user)
            deleted += 1
            total_unlinked += res.get("unlinked_products", 0)
            results.append({"id": cid, "ok": True, "unlinked_products": res.get("unlinked_products", 0)})
        except HTTPException as e:
            results.append({"id": cid, "ok": False, "error": e.detail})
        except Exception as e:  # never let one bad row abort the batch
            results.append({"id": cid, "ok": False, "error": str(e)})

    return {
        "ok": deleted == len(ids),
        "requested": len(ids),
        "deleted": deleted,
        "failed": len(ids) - deleted,
        "unlinked_products": total_unlinked,
        "results": results,
    }


@router.delete("/categories/{item_id}")
async def delete_category(
    item_id: str,
    force: bool = Query(False, description="Unlink linked products and delete anyway."),
    user: dict = Depends(require_admin),
):
    """Soft delete a category.

    By default the delete is blocked when products are still linked, so the
    operator notices and reassigns them. Pass ``force=true`` to unlink those
    products (clear their ``category``/``category_id``) and delete anyway — this
    keeps the data consistent instead of leaving products pointing at a deleted
    category. Sub-categories still block the delete regardless of ``force``.
    """
    force_val = False if not isinstance(force, bool) else force
    return await _delete_one_category(item_id, force_val, user)


# ───────────────────────── Bulk import (Excel / CSV) ─────────────────────────

def _norm_header(h) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(h or "").strip().lower()).strip("_")


# Accept friendly headers → internal fields.
_HEADER_MAP = {
    "name": "name", "category": "name", "category_name": "name",
    "code": "code", "category_code": "code",
    "description": "description", "desc": "description",
    "parent": "parent_name", "parent_category": "parent_name", "parent_name": "parent_name",
    "status": "status",
    "display_order": "display_order", "order": "display_order", "sort_order": "display_order",
}


def _rows_from_csv(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def _rows_from_xlsx(raw: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for r in rows_iter:
        if all(c is None for c in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out


@router.post("/categories/import")
async def import_categories(
    file: UploadFile = File(...),
    commit: bool = Query(True),
    user: dict = Depends(get_current_user),
):
    """Bulk-create categories from an uploaded Excel (.xlsx) or CSV file.

    Recognised columns (case/spacing-insensitive): Name (required), Code,
    Description, Parent Category, Status, Display Order. Parent is matched by
    name. With commit=false the file is validated and a dry-run report returned.
    """
    fname = (file.filename or "").lower()
    raw = await file.read()
    try:
        rows = _rows_from_xlsx(raw) if fname.endswith((".xlsx", ".xlsm")) else _rows_from_csv(raw)
    except Exception as e:  # malformed file
        raise HTTPException(400, f"Could not read the file: {e}")

    # Existing + within-file names for uniqueness and parent resolution.
    existing = {
        (c["name"] or "").strip().lower(): c["id"]
        async for c in db[COLL].find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1})
    }
    seen: set[str] = set()
    valid: list[dict] = []
    errors: list[dict] = []

    for i, raw_row in enumerate(rows):
        row = {}
        for k, v in raw_row.items():
            mapped = _HEADER_MAP.get(_norm_header(k))
            if mapped:
                row[mapped] = v
        name = str(row.get("name") or "").strip()
        if not name:
            errors.append({"row": i + 2, "error": "Missing category name"})
            continue
        key = name.lower()
        if key in existing or key in seen:
            errors.append({"row": i + 2, "name": name, "error": "Duplicate category name"})
            continue
        status = str(row.get("status") or "Active").strip().title()
        if status not in ("Active", "Inactive"):
            status = "Active"
        parent_id = None
        parent_name = str(row.get("parent_name") or "").strip()
        if parent_name:
            parent_id = existing.get(parent_name.lower())
            if not parent_id:
                errors.append({"row": i + 2, "name": name, "error": f"Parent '{parent_name}' not found"})
                continue
        try:
            order = int(float(row.get("display_order") or 0))
        except (TypeError, ValueError):
            order = 0
        valid.append({
            "name": name,
            "code": (str(row.get("code")).strip() if row.get("code") else None),
            "description": (str(row.get("description")).strip() if row.get("description") else None),
            "parent_id": parent_id,
            "status": status,
            "display_order": order,
        })
        seen.add(key)

    committed = 0
    if commit:
        for v in valid:
            v["code"] = await _unique_code((v["code"] or _slug_code(v["name"])))
            doc = {**v, "id": new_id(), "is_deleted": False,
                   "created_at": now_iso(), "updated_at": now_iso()}
            await db[COLL].insert_one(doc)
            committed += 1

    return {
        "total_rows": len(rows),
        "valid_rows": len(valid),
        "error_rows": len(errors),
        "errors": errors,
        "committed": committed,
        "dry_run": not commit,
    }
