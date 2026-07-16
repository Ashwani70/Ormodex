"""MIS Reports Router.

Advanced analytics, KPI dashboards, Excel/PDF export.
"""
import io
from datetime import date, timedelta
from typing import Optional, Any

from fastapi import APIRouter, Depends, HTTPException, Response

from core import cache
from core.auth_utils import get_current_user, is_admin_role
from core.db import db

router = APIRouter(prefix="/mis", tags=["MIS Reports"])


def _require_mis(user: dict):
    if is_admin_role(user.get("role")):
        return user
    perms = user.get("module_permissions", [])
    if not any(p in perms for p in ("mis_reports", "accounting", "reports")):
        raise HTTPException(403, "MIS Reports module access required")
    return user


# ─────────────────────────── Main KPI Dashboard ───────────────────────────

@router.get("/dashboard")
async def mis_dashboard(user=Depends(get_current_user)):
    _require_mis(user)
    return await cache.get_or_set(
        "mis:dashboard", cache.TTL_DASHBOARD, _compute_mis_dashboard
    )


async def _compute_mis_dashboard() -> dict:
    from sqlalchemy import text
    from core.db import get_session

    today = date.today()
    month_start = today.replace(day=1).isoformat()
    prev_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1).isoformat()
    prev_month_end = (today.replace(day=1) - timedelta(days=1)).isoformat()
    six_months_ago = (today - timedelta(days=180)).isoformat()

    # Single round-trip: run all 6 queries as one multi-statement batch via UNION ALL
    # so the Mumbai pooler pays one RTT instead of six (~30ms vs ~180ms+).
    COMBINED_SQL = text("""
        SELECT 'counts' AS q,
               (SELECT count(*) FROM products)::text AS a,
               (SELECT count(*) FROM products WHERE quantity <= low_stock_threshold)::text AS b,
               (SELECT count(*) FROM customers)::text AS c,
               (SELECT count(*) FROM vendors)::text AS d,
               (SELECT COALESCE(SUM(total - paid_amount), 0) FROM invoices WHERE status IN ('UNPAID','PARTIAL'))::text AS e,
               NULL::text AS f, NULL::text AS g, NULL::text AS h, NULL::text AS i, NULL::jsonb AS j
        UNION ALL
        SELECT 'invoice' AS q,
               i.total::text, i.paid_amount::text, i.created_at::text, i.status,
               cu.name, NULL, NULL, NULL, NULL, i.lines
        FROM invoices i
        LEFT JOIN customers cu ON i.customer_id = cu.id
        WHERE i.created_at >= :six_months_ago
        UNION ALL
        SELECT 'expense' AS q,
               e.amount::text, e.status, e.expense_date::text,
               NULL, NULL, NULL, NULL, NULL, NULL, NULL
        FROM expense_entries e
        WHERE e.expense_date >= :six_months_ago
        UNION ALL
        SELECT 'purchase' AS q,
               po.total_amount::text, po.created_at::text,
               NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL
        FROM purchase_orders po
        WHERE po.created_at >= :six_months_ago
    """)

    async with get_session() as session:
        rows = (await session.execute(COMBINED_SQL, {"six_months_ago": six_months_ago})).fetchall()

    product_count = low_stock_count = customer_count = supplier_count = 0
    receivables = 0.0
    invoices: list[dict] = []
    expenses: list[dict] = []
    purchases: list[dict] = []

    for row in rows:
        q = row[0]
        if q == "counts":
            product_count = int(row[1] or 0)
            low_stock_count = int(row[2] or 0)
            customer_count = int(row[3] or 0)
            supplier_count = int(row[4] or 0)
            receivables = round(float(row[5] or 0.0), 2)
        elif q == "invoice":
            raw_ts = row[3]
            invoices.append({
                "total": float(row[1] or 0.0),
                "payment_received": float(row[2] or 0.0),
                "created_at": raw_ts if isinstance(raw_ts, str) else (raw_ts.isoformat() if raw_ts else ""),
                "status": row[4],
                "customer_name": row[5],
                "items": row[10],
            })
        elif q == "expense":
            raw_dt = row[3]
            expenses.append({
                "amount": float(row[1] or 0.0),
                "status": row[2],
                "date": raw_dt if isinstance(raw_dt, str) else (raw_dt.isoformat() if raw_dt else ""),
            })
        elif q == "purchase":
            raw_ts = row[2]
            purchases.append({
                "total": float(row[1] or 0.0),
                "created_at": raw_ts if isinstance(raw_ts, str) else (raw_ts.isoformat() if raw_ts else ""),
            })

    def _f(v) -> float:
        try:
            return float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _cre(row) -> str:
        return str(row.get("created_at") or "")

    # Current vs previous month revenue (from the single invoices fetch).
    current_revenue = sum(_f(i.get("total")) for i in invoices if _cre(i) >= month_start)
    sales_count = sum(1 for i in invoices if _cre(i) >= month_start)
    prev_revenue = sum(
        _f(i.get("total")) for i in invoices
        if prev_month_start <= _cre(i) <= prev_month_end
    )
    # Guard the no-baseline case: when last month had zero revenue, a literal
    # %-change is undefined. The old `/max(prev,1)` turned ₹2.1L into a nonsense
    # "21,263,600%". Report 0 when there's no prior month to compare against.
    revenue_growth = (
        round(((current_revenue - prev_revenue) / prev_revenue) * 100, 1)
        if prev_revenue else 0.0
    )

    purchase_total = sum(_f(p.get("total")) for p in purchases if str(p.get("created_at") or "") >= month_start)

    expense_total = sum(
        _f(e.get("amount")) for e in expenses
        if str(e.get("date") or "") >= month_start and e.get("status") == "APPROVED"
    )

    # Monthly trend (last 6 months) — bucket the single fetches by month.
    trend_data = []
    for i in range(5, -1, -1):
        m_date = today.replace(day=1) - timedelta(days=i * 30)
        m_start = m_date.replace(day=1).isoformat()
        m_end = (m_date.replace(day=28) + timedelta(days=4)).replace(day=1).isoformat()
        label = m_date.strftime("%b %Y")
        m_rev = sum(_f(inv.get("total")) for inv in invoices if m_start <= _cre(inv) < m_end)
        m_exp = sum(
            _f(e.get("amount")) for e in expenses
            if m_start <= str(e.get("date") or "") < m_end and e.get("status") == "APPROVED"
        )
        trend_data.append({
            "month": label,
            "revenue": round(m_rev, 2),
            "expenses": round(m_exp, 2),
            "profit": round(m_rev - m_exp, 2),
        })

    # Top products by revenue (explode invoice line items in Python).
    prod_rev: dict[str, dict] = {}
    for inv in invoices:
        for it in (inv.get("items") or []):
            name = it.get("product_name") or "Unknown"
            d = prod_rev.setdefault(name, {"_id": name, "total_revenue": 0.0, "units_sold": 0.0})
            d["total_revenue"] += _f(it.get("quantity")) * _f(it.get("unit_price"))
            d["units_sold"] += _f(it.get("quantity"))
    top_products = sorted(prod_rev.values(), key=lambda d: d["total_revenue"], reverse=True)[:5]

    # Top customers by revenue.
    cust_rev: dict[str, dict] = {}
    for inv in invoices:
        name = inv.get("customer_name") or "Unknown"
        d = cust_rev.setdefault(name, {"_id": name, "total": 0.0, "count": 0})
        d["total"] += _f(inv.get("total"))
        d["count"] += 1
    top_customers = sorted(cust_rev.values(), key=lambda d: d["total"], reverse=True)[:5]

    return {
        "kpis": {
            "current_month_revenue": round(current_revenue, 2),
            "revenue_growth_pct": revenue_growth,
            "sales_count": sales_count,
            "purchase_total": round(purchase_total, 2),
            "expense_total": round(expense_total, 2),
            "gross_profit": round(current_revenue - purchase_total - expense_total, 2),
            "receivables_outstanding": round(receivables, 2),
            "customer_count": customer_count,
            "supplier_count": supplier_count,
            "product_count": product_count,
            "low_stock_count": low_stock_count,
        },
        "monthly_trend": trend_data,
        "top_products": top_products,
        "top_customers": top_customers,
    }


# ─────────────────────────── Sales Analysis ───────────────────────────

@router.get("/sales-analysis")
async def sales_analysis(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    group_by: str = "month",  # month | customer | product
    user=Depends(get_current_user)
):
    _require_mis(user)
    q: dict[str, Any] = {}
    if from_date or to_date:
        q["created_at"] = {}
        if from_date:
            q["created_at"]["$gte"] = from_date
        if to_date:
            q["created_at"]["$lte"] = to_date

    pipeline: list[dict[str, Any]] = []
    if group_by == "customer":
        pipeline = [
            {"$match": q},
            {"$group": {
                "_id": "$customer_name",
                "total_revenue": {"$sum": "$total"},
                "invoice_count": {"$sum": 1},
                "avg_invoice": {"$avg": "$total"},
            }},
            {"$sort": {"total_revenue": -1}},
        ]
    elif group_by == "product":
        pipeline = [
            {"$match": q},
            {"$unwind": "$items"},
            {"$group": {
                "_id": "$items.product_name",
                "total_revenue": {"$sum": {"$multiply": ["$items.quantity", "$items.unit_price"]}},
                "units_sold": {"$sum": "$items.quantity"},
            }},
            {"$sort": {"total_revenue": -1}},
        ]
    else:
        pipeline = [
            {"$match": q},
            {"$group": {
                "_id": {"$substr": ["$created_at", 0, 7]},
                "total_revenue": {"$sum": "$total"},
                "invoice_count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]

    results = await db.invoices.aggregate(pipeline).to_list(100)
    return {"group_by": group_by, "data": results}


# ─────────────────────────── Purchase Analysis ───────────────────────────

@router.get("/purchase-analysis")
async def purchase_analysis(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    group_by: str = "month",
    user=Depends(get_current_user)
):
    _require_mis(user)
    q: dict[str, Any] = {}
    if from_date or to_date:
        q["created_at"] = {}
        if from_date:
            q["created_at"]["$gte"] = from_date
        if to_date:
            q["created_at"]["$lte"] = to_date

    pipeline: list[dict[str, Any]] = []
    if group_by == "supplier":
        pipeline = [
            {"$match": q},
            {"$group": {
                "_id": "$supplier_name",
                "total_purchase": {"$sum": "$total"},
                "order_count": {"$sum": 1},
            }},
            {"$sort": {"total_purchase": -1}},
        ]
    elif group_by == "product":
        pipeline = [
            {"$match": q},
            {"$unwind": "$items"},
            {"$group": {
                "_id": "$items.product_name",
                "total_cost": {"$sum": {"$multiply": ["$items.quantity", "$items.unit_price"]}},
                "total_qty": {"$sum": "$items.quantity"},
            }},
            {"$sort": {"total_cost": -1}},
        ]
    else:
        pipeline = [
            {"$match": q},
            {"$group": {
                "_id": {"$substr": ["$created_at", 0, 7]},
                "total_purchase": {"$sum": "$total"},
                "order_count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]

    results = await db.purchase_orders.aggregate(pipeline).to_list(100)
    return {"group_by": group_by, "data": results}


# ─────────────────────────── Profitability Report ───────────────────────────

@router.get("/profitability")
async def profitability_report(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    _require_mis(user)
    q_sales: dict[str, Any] = {}
    q_exp: dict[str, Any] = {}
    if from_date or to_date:
        if from_date:
            q_sales.setdefault("created_at", {})["$gte"] = from_date
            q_exp.setdefault("date", {})["$gte"] = from_date
        if to_date:
            q_sales.setdefault("created_at", {})["$lte"] = to_date
            q_exp.setdefault("date", {})["$lte"] = to_date

    q_exp["status"] = "APPROVED"

    # NOTE: these three aggregates are logically independent, but they can't
    # be run concurrently via asyncio.gather() — this app binds ONE shared
    # AsyncSession per HTTP request (core/db.py's request-scoped ContextVar,
    # set up in server.py's request_db_session middleware) specifically so
    # every db.* call in a request reuses one pooled connection. Concurrent
    # queries against that same session raise SQLAlchemy's
    # IllegalStateChangeError (confirmed live — this was tried and reverted).
    # The dashboard summary endpoint's parallel queries only work because
    # they open their OWN separate AsyncSessionLocal() sessions directly via
    # raw SQL, bypassing this shared session entirely — a much bigger change
    # than is warranted here for 3 lightweight aggregates.
    sales_agg = await db.invoices.aggregate([
        {"$match": q_sales},
        {"$group": {"_id": None, "revenue": {"$sum": "$total"}}},
    ]).to_list(1)
    purchase_agg = await db.purchase_orders.aggregate([
        {"$match": q_sales},
        {"$group": {"_id": None, "cogs": {"$sum": "$total"}}},
    ]).to_list(1)
    expense_agg = await db.expense_entries.aggregate([
        {"$match": q_exp},
        {"$group": {
            "_id": "$category",
            "total": {"$sum": "$amount"},
        }},
    ]).to_list(50)
    revenue = sales_agg[0]["revenue"] if sales_agg else 0
    cogs = purchase_agg[0]["cogs"] if purchase_agg else 0
    total_expenses = sum(e["total"] for e in expense_agg)

    gross_profit = revenue - cogs
    net_profit = gross_profit - total_expenses
    gross_margin = round((gross_profit / max(revenue, 1)) * 100, 1)
    net_margin = round((net_profit / max(revenue, 1)) * 100, 1)

    return {
        "revenue": round(revenue, 2),
        "cogs": round(cogs, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_margin_pct": gross_margin,
        "operating_expenses": round(total_expenses, 2),
        "expense_breakdown": expense_agg,
        "net_profit": round(net_profit, 2),
        "net_margin_pct": net_margin,
        "from_date": from_date,
        "to_date": to_date,
    }


# ─────────────────────────── Excel Export ───────────────────────────

@router.get("/export/sales-excel")
async def export_sales_excel(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    _require_mis(user)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    q: dict[str, Any] = {}
    if from_date:
        q.setdefault("created_at", {})["$gte"] = from_date
    if to_date:
        q.setdefault("created_at", {})["$lte"] = to_date

    invoices = await db.invoices.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(500, "Could not create active worksheet")
    ws.title = "Sales Report"

    # Header style
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFD700", bold=True)

    headers = ["Invoice No", "Date", "Customer", "Currency", "Total", "Paid", "Outstanding", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row_idx, inv in enumerate(invoices, 2):
        total = inv.get("total", 0)
        paid = inv.get("payment_received", 0)
        ws.cell(row=row_idx, column=1, value=inv.get("invoice_number", ""))
        ws.cell(row=row_idx, column=2, value=inv.get("created_at", "")[:10])
        ws.cell(row=row_idx, column=3, value=inv.get("customer_name", ""))
        ws.cell(row=row_idx, column=4, value=inv.get("currency", "INR"))
        ws.cell(row=row_idx, column=5, value=round(total, 2))
        ws.cell(row=row_idx, column=6, value=round(paid, 2))
        ws.cell(row=row_idx, column=7, value=round(total - paid, 2))
        ws.cell(row=row_idx, column=8, value=inv.get("status", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"sales_report_{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/expense-excel")
async def export_expense_excel(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    _require_mis(user)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    q: dict[str, Any] = {"status": "APPROVED"}
    if from_date:
        q.setdefault("date", {})["$gte"] = from_date
    if to_date:
        q.setdefault("date", {})["$lte"] = to_date

    expenses = await db.expense_entries.find(q, {"_id": 0}).sort("date", -1).to_list(2000)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(500, "Could not create active worksheet")
    ws.title = "Expense Report"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFD700", bold=True)

    headers = ["Date", "Category", "Department", "Description", "Amount", "Payment Mode", "Status", "Approved By"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 18

    for i, e in enumerate(expenses, 2):
        ws.cell(row=i, column=1, value=e.get("date", ""))
        ws.cell(row=i, column=2, value=e.get("category", ""))
        ws.cell(row=i, column=3, value=e.get("department", ""))
        ws.cell(row=i, column=4, value=e.get("description", ""))
        ws.cell(row=i, column=5, value=round(e.get("amount", 0), 2))
        ws.cell(row=i, column=6, value=e.get("payment_mode", ""))
        ws.cell(row=i, column=7, value=e.get("status", ""))
        ws.cell(row=i, column=8, value=e.get("approved_by_name", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"expense_report_{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
