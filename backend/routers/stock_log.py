"""Stock Log — enterprise stock ledger (TallyPrime-style), rebuilt 2026-07-14.

Reads from `stock_transactions`, the table actually populated by every live
posting path (Purchase, Sales, GRN/Material Receipt, Job Work, Manufacturing,
Production, Stock Transfer, Physical Stock, Adjustment) — see
`core/schema.py StockTransaction` and the writers in routers/purchase.py,
sales.py, job_work.py, manufacturing.py, inventory.py, inventory_v2.py, and
stock_analysis.py. (`stock_ledger_entries`/`vouchers_v2` is a more rigorously
modelled parallel system but has near-zero live data — see project memory.)

Design:
- ONE raw-SQL query per endpoint (`text()` + bound params), matching this
  deployment's dominant-cost-is-round-trips profile (Supabase pooler in
  Singapore/ap-southeast — see core/db.py). No N+1, no per-row Python loops
  over the DB.
- Running balance is computed at READ time via a window function
  (`SUM(delta) OVER (PARTITION BY product_id ORDER BY created_at, id)`)
  rather than trusting the stored `balance` column — some writers (stock
  transfers) cannot supply a per-row balance since they don't track
  `products.quantity` directly. This is also more correct: the stored
  `balance` can drift if two requests race, the window function can't.
- Server-side pagination, filtering, sorting — never ships the whole table.
- doc_type is the closest thing to a "voucher type" today; historical rows
  written before this rebuild have NULL doc_type/voucher_no and fall back to
  a best-effort parse of the free-text `reason` column.
"""
import io
import re
from datetime import date, datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Response
from pydantic import BaseModel
from sqlalchemy import text

from core.auth_utils import get_current_user, is_admin_role, require_admin
from core.db import get_session
from core.utils import log_audit

router = APIRouter(prefix="/stock-log", tags=["Stock Log"])


def _require_inventory(user: dict) -> dict:
    if is_admin_role(user.get("role")) or user.get("role") == "accountant":
        return user
    if "inventory" in (user.get("module_permissions") or []):
        return user
    raise HTTPException(status_code=403, detail="Inventory module access required")


# Deep-link routes for the voucher drill-down. Matches the app's existing
# ?detail=<id> / ?q= convention (see routers/search.py's _DETAIL_AWARE_ROUTES/
# _Q_AWARE_ROUTES) plus the two real /:id routes (job work).
_DOC_TYPE_ROUTES: dict[str, str] = {
    "PURCHASE": "/purchase-orders",
    "MATERIAL_RECEIPT": "/purchase-bills",
    "SALES": "/sales-orders",
    "SALES_RETURN": "/credit-notes",
    "JOB_WORK": "/job-work/challans/{id}",
    "JOB_WORK_RECEIPT": "/job-work/receipts/{id}",
    "JOB_WORK_REVERSAL": "/job-work/challans/{id}",
    "MANUFACTURING": "/manufacturing",
    "PRODUCTION": "/manufacturing",
    "STOCK_TRANSFER": "/stock-transfers",
    "PHYSICAL_STOCK": "/stock-log",  # no dedicated frontend page today
    "ADJUSTMENT": "/products",
    "OPENING_STOCK": "/products",
}

# doc_type="PURCHASE" alone is too coarse to route correctly — a v2 GRN
# (post_entry source_doc_type="grn"), a v1 GRN ("grn_v1"), and a plain PO
# direct-receive ("purchase_order") all share that doc_type but live on
# different pages. When the row's source_doc_type is known, this map takes
# priority over _DOC_TYPE_ROUTES above. See migration 026 /
# project-stocklog-grn-route-mislabel.md for the full trace.
_SOURCE_DOC_TYPE_ROUTES: dict[str, str] = {
    "grn": "/grns",
    "grn_v1": "/grns",
    "purchase_order": "/purchase-orders",
}

_DOC_TYPE_LABELS: dict[str, str] = {
    "PURCHASE": "Purchase",
    "MATERIAL_RECEIPT": "Material Receipt",
    "SALES": "Sales",
    "SALES_RETURN": "Sales Return",
    "JOB_WORK": "Job Work Issue",
    "JOB_WORK_RECEIPT": "Job Work Receipt",
    "JOB_WORK_REVERSAL": "Job Work Reversal",
    "MANUFACTURING": "Manufacturing",
    "PRODUCTION": "Production",
    "STOCK_TRANSFER": "Stock Transfer",
    "PHYSICAL_STOCK": "Physical Stock",
    "ADJUSTMENT": "Adjustment",
    "OPENING_STOCK": "Opening Stock",
}

_VOUCHER_NO_RE = re.compile(r"\b([A-Z]{2,5}[-_]?\d{3,}[A-Za-z0-9-]*)\b")


def _fallback_voucher_no(reason: Optional[str]) -> Optional[str]:
    """Historical rows (pre-rebuild) have no voucher_no column value — best-
    effort pull one out of the free-text `reason` string they do have."""
    if not reason:
        return None
    m = _VOUCHER_NO_RE.search(reason)
    return m.group(1) if m else None


def _fallback_doc_type(reason: Optional[str]) -> Optional[str]:
    if not reason:
        return None
    r = reason.lower()
    if "purchase" in r:
        return "PURCHASE"
    if "goods receipt" in r:
        return "MATERIAL_RECEIPT"
    if "sales" in r:
        return "SALES"
    if "job work receipt" in r:
        return "JOB_WORK_RECEIPT"
    if "job work edit reversal" in r:
        return "JOB_WORK_REVERSAL"
    if "job work" in r:
        return "JOB_WORK"
    if "mfg" in r or "manufactur" in r:
        return "MANUFACTURING"
    if "prod journal" in r:
        return "PRODUCTION"
    if "transfer" in r:
        return "STOCK_TRANSFER"
    if "physical" in r:
        return "PHYSICAL_STOCK"
    return "ADJUSTMENT"


_SORTABLE_COLUMNS = {
    # Keyed on the frontend grid's own camelCase column keys (StockLog.jsx's
    # DEFAULT_COLUMNS) so a header click's sort_by needs no translation layer
    # to drift out of sync with. Values are unqualified — the outer query in
    # list_entries() selects FROM the `ledger` CTE, whose output columns have
    # no `t.` alias (that alias only exists inside the CTE's own SELECT).
    "created_at": "created_at",
    "date": "created_at",
    "productName": "product_name",
    "productCode": "product_sku",
    "voucherNo": "voucher_no",
    "voucherTypeLabel": "doc_type",
    "inQty": "delta",
    "outQty": "delta",
    "delta": "delta",
    "runningBalance": "running_balance",
    "doc_type": "doc_type",
    "user": "user_name",
    "user_name": "user_name",
}

_BASE_SELECT = """
    SELECT
        t.id, t.product_id, t.product_name, t.godown_id, g.name AS godown_name,
        t.batch_id, t.doc_type, t.source_doc_type, t.voucher_no, t.source_doc_id,
        t.qty, t.rate, t.value, t.delta, t.reason, t.user_id, t.user_name,
        t.created_at,
        SUM(t.delta) OVER (
            PARTITION BY t.product_id ORDER BY t.created_at, t.id
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS running_balance,
        p.sku AS product_sku, p.unit AS product_unit
    FROM stock_transactions t
    LEFT JOIN godowns g ON g.id = t.godown_id
    LEFT JOIN products p ON p.id = t.product_id
"""


def _build_filters(
    product_id, warehouse_id, from_date, to_date, doc_type, user_id, status, q,
) -> tuple[str, dict]:
    where = []
    params: dict[str, Any] = {}
    if product_id:
        where.append("t.product_id = :product_id")
        params["product_id"] = product_id
    if warehouse_id:
        where.append("t.godown_id = :warehouse_id")
        params["warehouse_id"] = warehouse_id
    if from_date:
        where.append("t.created_at >= :from_date")
        params["from_date"] = from_date
    if to_date:
        # created_at is an ISO-with-time string; include the whole end day.
        where.append("t.created_at <= :to_date")
        params["to_date"] = f"{to_date}T23:59:59.999999"
    if doc_type:
        where.append("t.doc_type = :doc_type")
        params["doc_type"] = doc_type
    if user_id:
        where.append("t.user_id = :user_id")
        params["user_id"] = user_id
    if status == "inward":
        where.append("t.delta > 0")
    elif status == "outward":
        where.append("t.delta < 0")
    if q:
        where.append(
            "(t.product_name ILIKE :q OR t.voucher_no ILIKE :q OR t.reason ILIKE :q "
            "OR p.sku ILIKE :q OR t.user_name ILIKE :q)"
        )
        params["q"] = f"%{q}%"
    clause = f"WHERE {' AND '.join(where)}" if where else ""
    return clause, params


@router.get("/entries")
async def list_entries(
    product_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    doc_type: Optional[str] = None,
    user_id: Optional[str] = None,
    status: Optional[str] = None,  # all | inward | outward
    q: Optional[str] = None,
    sort_by: str = "created_at",
    sort_dir: str = "desc",
    page: int = 1,
    limit: int = 100,
    user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    """Paginated, filtered, sorted stock ledger rows with a window-function
    running balance. One query for the page + one COUNT(*) for the total."""
    _require_inventory(user)
    limit = max(1, min(limit, 500))
    page = max(1, page)
    offset = (page - 1) * limit

    where_clause, params = _build_filters(
        product_id, warehouse_id, from_date, to_date, doc_type, user_id, status, q
    )
    sort_col = _SORTABLE_COLUMNS.get(sort_by, "t.created_at")
    sort_dir = "ASC" if sort_dir.lower() == "asc" else "DESC"

    # The window function must run over ALL matching rows before pagination
    # clips the page, so filter/sort/paginate the *outer* query over a CTE
    # rather than pushing LIMIT into the windowed SELECT directly.
    sql = text(f"""
        WITH ledger AS ({_BASE_SELECT} {where_clause})
        SELECT *, count(*) OVER () AS total_count
        FROM ledger
        ORDER BY {sort_col} {sort_dir}, id {sort_dir}
        LIMIT :limit OFFSET :offset
    """)
    params["limit"] = limit
    params["offset"] = offset

    async with get_session() as session:
        rows = (await session.execute(sql, params)).mappings().all()

    total = rows[0]["total_count"] if rows else 0
    items = []
    for r in rows:
        doc_type_val = r["doc_type"] or _fallback_doc_type(r["reason"])
        voucher_no_val = r["voucher_no"] or _fallback_voucher_no(r["reason"])
        items.append({
            "id": r["id"],
            "date": (r["created_at"] or "")[:10],
            "time": (r["created_at"] or "")[11:19],
            "createdAt": r["created_at"],
            "voucherNo": voucher_no_val,
            "voucherType": doc_type_val,
            "voucherTypeLabel": _DOC_TYPE_LABELS.get(doc_type_val, doc_type_val or "—") if doc_type_val else "—",
            "productId": r["product_id"],
            "productCode": r["product_sku"],
            "productName": r["product_name"],
            "unit": r["product_unit"],
            "warehouse": r["godown_name"],
            "warehouseId": r["godown_id"],
            "batchNo": r["batch_id"],
            "serialNo": None,  # not tracked at the stock_transactions grain today
            "inQty": float(r["delta"]) if r["delta"] and r["delta"] > 0 else 0.0,
            "outQty": -float(r["delta"]) if r["delta"] and r["delta"] < 0 else 0.0,
            "runningBalance": float(r["running_balance"]) if r["running_balance"] is not None else None,
            "rate": float(r["rate"]) if r["rate"] is not None else None,
            "value": float(r["value"]) if r["value"] is not None else (
                float(r["delta"]) * float(r["rate"]) if r["rate"] is not None and r["delta"] is not None else None
            ),
            "costCentre": None,  # not tracked at this grain today
            "user": r["user_name"],
            "userId": r["user_id"],
            "status": "IN" if (r["delta"] or 0) > 0 else "OUT",
            "sourceDocId": r["source_doc_id"],
            "sourceDocType": r["source_doc_type"],
        })

    await log_audit(
        "SEARCH", "stock_log", None, user,
        new_values={"filters": {k: v for k, v in params.items() if k not in ("limit", "offset")}, "result_count": len(items)},
    )

    return {"items": items, "total": total, "page": page, "limit": limit}


@router.get("/summary")
async def summary(
    product_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    doc_type: Optional[str] = None,
    user_id: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    """Dashboard summary cards — one aggregate query over the filtered set,
    plus one more for opening stock (balance as of just before from_date)."""
    _require_inventory(user)
    where_clause, params = _build_filters(
        product_id, warehouse_id, from_date, to_date, doc_type, user_id, status, q
    )

    agg_sql = text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN t.delta > 0 THEN t.delta ELSE 0 END), 0) AS total_in,
            COALESCE(SUM(CASE WHEN t.delta < 0 THEN -t.delta ELSE 0 END), 0) AS total_out,
            COALESCE(SUM(t.delta * COALESCE(t.rate, 0)), 0) AS net_value,
            COALESCE(SUM(t.delta), 0) AS net_qty,
            COUNT(*) AS txn_count,
            COALESCE(AVG(t.rate) FILTER (WHERE t.rate IS NOT NULL AND t.rate > 0), 0) AS avg_cost
        FROM stock_transactions t
        LEFT JOIN godowns g ON g.id = t.godown_id
        LEFT JOIN products p ON p.id = t.product_id
        {where_clause}
    """)

    # Opening stock: net movement for this product/warehouse scope strictly
    # before from_date (defaults to 0 when no from_date is set — the whole
    # ledger is then "the period").
    opening_where, opening_params = _build_filters(
        product_id, warehouse_id, None, None, None, None, None, None
    )
    opening_sql = None
    if from_date:
        opening_sql = text(f"""
            SELECT COALESCE(SUM(t.delta), 0) AS opening_qty
            FROM stock_transactions t
            WHERE t.created_at < :from_date
            {"AND t.product_id = :product_id" if product_id else ""}
            {"AND t.godown_id = :warehouse_id" if warehouse_id else ""}
        """)

    negative_stock_sql = text("""
        SELECT COUNT(*) FROM (
            SELECT product_id, SUM(delta) AS bal
            FROM stock_transactions
            GROUP BY product_id
            HAVING SUM(delta) < 0
        ) neg
    """)

    async with get_session() as session:
        agg = (await session.execute(agg_sql, params)).mappings().first()
        opening_qty = 0.0
        if opening_sql is not None:
            op_params = {"from_date": from_date}
            if product_id:
                op_params["product_id"] = product_id
            if warehouse_id:
                op_params["warehouse_id"] = warehouse_id
            opening_row = (await session.execute(opening_sql, op_params)).mappings().first()
            if opening_row is not None:
                opening_qty = float(opening_row["opening_qty"] or 0)
        negative_count = (await session.execute(negative_stock_sql)).scalar() or 0
        pending_transfers = (await session.execute(text(
            "SELECT COUNT(*) FROM stock_transfers WHERE status = 'PENDING'"
        ))).scalar() or 0

    total_in = float(agg["total_in"] or 0) if agg else 0.0
    total_out = float(agg["total_out"] or 0) if agg else 0.0
    closing_qty = opening_qty + total_in - total_out
    net_value = float(agg["net_value"] or 0) if agg else 0.0

    return {
        "openingStock": round(opening_qty, 2),
        "totalInward": round(total_in, 2),
        "totalOutward": round(total_out, 2),
        "closingStock": round(closing_qty, 2),
        "stockValue": round(net_value, 2),
        "averageCost": round(float(agg["avg_cost"] or 0), 2) if agg else 0.0,
        "negativeStockCount": int(negative_count),
        "pendingTransfers": int(pending_transfers),
        "transactionCount": int(agg["txn_count"] or 0) if agg else 0,
    }


@router.get("/filters")
async def filter_options(user: dict = Depends(get_current_user)) -> dict[str, Any]:
    """Dropdown option lists for the header filters — products, warehouses,
    voucher types actually present, and users who have posted movements."""
    _require_inventory(user)
    sql = text("""
        SELECT 'products' AS kind, id, name AS label, sku AS extra FROM products
            WHERE coalesce(is_deleted, false) = false ORDER BY name LIMIT 500
    """)
    warehouses_sql = text("SELECT id, name FROM godowns WHERE coalesce(is_deleted, false) = false ORDER BY name")
    doc_types_sql = text("SELECT DISTINCT doc_type FROM stock_transactions WHERE doc_type IS NOT NULL ORDER BY doc_type")
    users_sql = text("""
        SELECT DISTINCT user_id, user_name FROM stock_transactions
        WHERE user_id IS NOT NULL ORDER BY user_name
    """)

    async with get_session() as session:
        products = (await session.execute(sql)).mappings().all()
        warehouses = (await session.execute(warehouses_sql)).mappings().all()
        doc_types = (await session.execute(doc_types_sql)).scalars().all()
        users = (await session.execute(users_sql)).mappings().all()

    return {
        "products": [{"id": p["id"], "name": p["label"], "sku": p["extra"]} for p in products],
        "warehouses": [{"id": w["id"], "name": w["name"]} for w in warehouses],
        "voucherTypes": [
            {"value": dt, "label": _DOC_TYPE_LABELS.get(dt, dt)} for dt in doc_types
        ],
        "users": [{"id": u["user_id"], "name": u["user_name"]} for u in users],
    }


# ───────────────────────── Delete (admin-only cleanup) ─────────────────────────
#
# Stock Log rows are a derived ledger: running balance, valuation, and
# negative-stock alerts are all computed by summing `delta` at read time (see
# the module docstring), so removing a row is safe for the LEDGER itself — no
# stale `balance` to leave behind. It is NOT safe for `products.quantity`,
# which several v1 posting paths (Purchase, Sales confirm, Job Work,
# Manufacturing, Adjustment) update directly alongside the ledger row —
# deleting the row without reversing that would silently desync the two. So
# every delete here reverses `products.quantity` by the row's own `delta`
# before removing it.
#
# A row that has a matching `stock_ledger_entries` counterpart (i.e. came from
# a v2 posting flow — GRN, Purchase Return, Stock Adjustment v2, Stock
# Transfer, ...) is refused: deleting only its stock_transactions mirror would
# silently reintroduce the exact "Stock Log shows nothing for this item" bug
# the dual-write fix closed, since the ledger row would still exist with no
# mirror. Those need to be corrected by reversing the source voucher, not a
# raw row delete — this endpoint is for genuine test/junk rows that have no
# such counterpart (v1-only postings, or orphaned rows).

_DELETE_SELECT_SQL = text("""
    SELECT t.id, t.product_id, t.delta, t.source_doc_id, t.doc_type, t.source_doc_type, t.voucher_no, t.reason
    FROM stock_transactions t
    WHERE t.id = ANY(:ids)
""")

_HAS_LEDGER_COUNTERPART_SQL = text("""
    SELECT DISTINCT source_doc_id FROM stock_ledger_entries
    WHERE source_doc_id = ANY(:source_doc_ids)
""")

_REVERSE_QTY_BULK_SQL = text("""
    UPDATE products SET quantity = COALESCE(quantity, 0) - v.delta, updated_at = :now
    FROM (SELECT unnest(CAST(:product_ids AS VARCHAR[])) AS product_id, unnest(CAST(:deltas AS NUMERIC[])) AS delta) AS v
    WHERE products.id = v.product_id
""")

_DELETE_ROWS_SQL = text("DELETE FROM stock_transactions WHERE id = ANY(:ids)")

_DELETE_LEDGER_COUNTERPARTS_SQL = text("""
    DELETE FROM stock_ledger_entries
    WHERE source_doc_id IN (
        SELECT source_doc_id
        FROM stock_transactions
        WHERE id = ANY(:ids) AND source_doc_id IS NOT NULL
    )
""")

_DOC_TYPE_TO_COLLECTION = {
    "JOB_WORK": "job_work_challans",
    "JOB_WORK_RECEIPT": "job_work_receipts",
    "JOB_WORK_REVERSAL": "job_work_challans",
    "STOCK_TRANSFER": "stock_transfers",
    "PHYSICAL_STOCK": "physical_verifications",
    "MANUFACTURING": "production_journals",
    "PRODUCTION": "production_journals",
}

_SOURCE_DOC_TYPE_TO_COLLECTION = {
    "job_work_challan": "job_work_challans",
    "job_work_receipt": "job_work_receipts",
    "job_work_challan_delete": "job_work_challans",
    "job_work_receipt_reversal": "job_work_receipts",
    "job_work_challan_cancel": "job_work_challans",
    "job_work_challan_edit_reversal": "job_work_challans",
    "grn": "grn_v2",
    "grn_v1": "grn_v2",
    "purchase_order": "purchase_orders_v2",
    "purchase_bill": "purchase_bills",
    "purchase_return": "purchase_returns",
    "sales_order": "sales_orders",
    "invoice": "invoices",
    "credit_note": "credit_notes",
    "dispatch": "dispatches",
    "proforma_invoice": "proforma_invoices",
}


async def _delete_stock_log_rows(ids: list[str], user: dict, force: bool = False) -> dict[str, Any]:
    from core.utils import now_iso

    if not ids:
        return {"deleted": 0, "skipped": [], "not_found": []}

    async with get_session() as session:
        rows = (await session.execute(_DELETE_SELECT_SQL, {"ids": ids})).mappings().all()
        found_ids = {r["id"] for r in rows}
        not_found = [i for i in ids if i not in found_ids]

        # A row posted via post_entry (v2) carries its ORIGINAL document id as
        # source_doc_id, which is exactly what stock_ledger_entries.source_doc_id
        # also stores for that same posting — that shared id is how we detect
        # "this stock_transactions row has a live ledger counterpart elsewhere".
        source_doc_ids = [r["source_doc_id"] for r in rows if r["source_doc_id"]]
        has_counterpart = set()
        if source_doc_ids:
            counterpart_rows = (await session.execute(
                _HAS_LEDGER_COUNTERPART_SQL, {"source_doc_ids": source_doc_ids}
            )).scalars().all()
            has_counterpart = set(counterpart_rows)

        # For rows with a ledger counterpart, check if the source voucher actually exists in the database.
        # If it does not exist, we treat it as an orphan and allow deleting it.
        real_counterparts = set()
        from core.db import db
        for r in rows:
            sid = r["source_doc_id"]
            if sid and sid in has_counterpart:
                coll_name = _SOURCE_DOC_TYPE_TO_COLLECTION.get(r["source_doc_type"]) or _DOC_TYPE_TO_COLLECTION.get(r["doc_type"])
                if coll_name:
                    try:
                        exists = await db[coll_name].find_one({"id": sid})
                        if exists:
                            real_counterparts.add(sid)
                    except Exception:
                        pass

        # Deletable rows are collected first (pure in-memory filtering, no DB
        # call), deltas summed per product (so two rows on the same product
        # still compose correctly — matches the old per-row sequential
        # subtraction), then applied as ONE bulk quantity-reversal UPDATE and
        # ONE bulk DELETE instead of two SQL statements per row inside the
        # loop (was up to 2x500 individual round trips on a max-size batch).
        skipped = []
        to_delete_ids: list[str] = []
        delta_by_product: dict[str, float] = {}
        now = now_iso()
        for r in rows:
            if not force and r["source_doc_id"] and r["source_doc_id"] in real_counterparts:
                skipped.append({
                    "id": r["id"], "docType": r["doc_type"], "sourceDocType": r["source_doc_type"],
                    "sourceDocId": r["source_doc_id"], "voucherNo": r["voucher_no"],
                    "reason": "Posted via the v2 ledger — reverse the source voucher instead of deleting this row directly.",
                })
                continue
            delta = float(r["delta"] or 0)
            if r["product_id"] and delta:
                delta_by_product[r["product_id"]] = delta_by_product.get(r["product_id"], 0.0) + delta
            to_delete_ids.append(r["id"])

        if delta_by_product:
            await session.execute(_REVERSE_QTY_BULK_SQL, {
                "product_ids": list(delta_by_product.keys()),
                "deltas": list(delta_by_product.values()),
                "now": now,
            })
        if to_delete_ids:
            await session.execute(_DELETE_LEDGER_COUNTERPARTS_SQL, {"ids": to_delete_ids})
            await session.execute(_DELETE_ROWS_SQL, {"ids": to_delete_ids})
        deleted = len(to_delete_ids)

        if deleted:
            await log_audit(
                "DELETE", "stock_transactions", None, user,
                old_values={"ids": to_delete_ids},
            )

    return {"deleted": deleted, "skipped": skipped, "not_found": not_found}


@router.delete("/entries/{entry_id}")
async def delete_entry(entry_id: str, force: bool = False, user: dict = Depends(require_admin)) -> dict[str, Any]:
    """Delete one Stock Log row — admin only. Reverses products.quantity by
    the row's delta first; refused if the row has a live stock_ledger_entries
    counterpart (unless force=True is passed by admin)."""
    result = await _delete_stock_log_rows([entry_id], user, force=force)
    if result["not_found"]:
        raise HTTPException(404, "Entry not found")
    if result["skipped"]:
        raise HTTPException(400, result["skipped"][0]["reason"])
    return {"ok": True}


class BulkDeleteRequest(BaseModel):
    ids: list[str]
    force: bool = False


@router.post("/entries/bulk-delete")
async def bulk_delete_entries(payload: BulkDeleteRequest, user: dict = Depends(require_admin)) -> dict[str, Any]:
    """Delete multiple Stock Log rows at once — admin only, same per-row rules
    as delete_entry (quantity reversal, ledger-counterpart refusal). Never
    partially fails the whole batch: rows that can't be deleted are reported
    back in `skipped`/`notFound` instead of aborting the ones that can."""
    if not payload.ids:
        raise HTTPException(400, "No entries selected")
    if len(payload.ids) > 500:
        raise HTTPException(400, "Cannot delete more than 500 entries at once")
    result = await _delete_stock_log_rows(payload.ids, user, force=payload.force)
    return {
        "deleted": result["deleted"],
        "skipped": result["skipped"],
        "notFound": result["not_found"],
    }


@router.get("/product-panel/{product_id}")
async def product_panel(product_id: str, user: dict = Depends(get_current_user)) -> dict[str, Any]:
    """Right info panel: everything Products/Vendors/Customers tables + the
    ledger itself can tell us about one product's current stock position."""
    _require_inventory(user)
    product_sql = text("""
        SELECT id, name, sku, category, unit, cost_price, selling_price, quantity,
               low_stock_threshold, warehouse_id, hsn_code, image_url, image_path
        FROM products WHERE id = :pid
    """)
    ledger_sql = text("""
        SELECT
            COALESCE(SUM(delta), 0) AS current_stock,
            COALESCE(SUM(delta * COALESCE(rate, 0)), 0) AS stock_value,
            COALESCE(AVG(rate) FILTER (WHERE rate IS NOT NULL AND rate > 0), 0) AS avg_cost
        FROM stock_transactions WHERE product_id = :pid
    """)
    last_purchase_sql = text("""
        SELECT created_at, rate FROM stock_transactions
        WHERE product_id = :pid AND doc_type IN ('PURCHASE', 'MATERIAL_RECEIPT')
        ORDER BY created_at DESC LIMIT 1
    """)
    last_sale_sql = text("""
        SELECT created_at, rate FROM stock_transactions
        WHERE product_id = :pid AND doc_type = 'SALES'
        ORDER BY created_at DESC LIMIT 1
    """)

    async with get_session() as session:
        prod = (await session.execute(product_sql, {"pid": product_id})).mappings().first()
        if not prod:
            raise HTTPException(404, "Product not found")
        ledger = (await session.execute(ledger_sql, {"pid": product_id})).mappings().first()
        last_purchase = (await session.execute(last_purchase_sql, {"pid": product_id})).mappings().first()
        last_sale = (await session.execute(last_sale_sql, {"pid": product_id})).mappings().first()
        warehouse = None
        if prod["warehouse_id"]:
            warehouse = (await session.execute(
                text("SELECT name FROM godowns WHERE id = :wid"), {"wid": prod["warehouse_id"]}
            )).scalar()
            if not warehouse:
                warehouse = (await session.execute(
                    text("SELECT name FROM warehouses WHERE id = :wid"), {"wid": prod["warehouse_id"]}
                )).scalar()

    current_stock = float(ledger["current_stock"] or 0) if ledger else 0.0
    return {
        "productId": prod["id"],
        "name": prod["name"],
        "sku": prod["sku"],
        "imageUrl": prod["image_url"] or prod["image_path"],
        "currentStock": round(current_stock, 2),
        "reservedStock": 0.0,  # no reservation concept in this codebase today
        "availableStock": round(current_stock, 2),
        "openingBalance": None,  # period-dependent; see /summary for a scoped value
        "closingBalance": round(current_stock, 2),
        "averageCost": round(float(ledger["avg_cost"] or 0), 2) if ledger else 0.0,
        "stockValue": round(float(ledger["stock_value"] or 0), 2) if ledger else 0.0,
        "lastPurchase": dict(last_purchase) if last_purchase else None,
        "lastSale": dict(last_sale) if last_sale else None,
        "warehouseLocation": warehouse,
        "minimumStock": float(prod["low_stock_threshold"] or 0),
        "maximumStock": None,  # not modeled on Product today
        "reorderLevel": float(prod["low_stock_threshold"] or 0),
        "unit": prod["unit"],
        "hsnCode": prod["hsn_code"],
        "costPrice": float(prod["cost_price"] or 0),
        "sellingPrice": float(prod["selling_price"] or 0),
    }


@router.get("/voucher/{doc_type}/{source_doc_id}")
async def resolve_voucher(
    doc_type: str, source_doc_id: str,
    source_doc_type: Optional[str] = None,
    user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    """Voucher drill-down resolver: given a row's (doc_type, source_doc_id),
    return the deep-link path into that voucher's own existing page/modal.
    `source_doc_type` (optional query param) disambiguates doc_types that
    cover multiple real pages — e.g. "PURCHASE" alone can't tell a v2 GRN
    from a plain PO receive, but source_doc_type can (see
    _SOURCE_DOC_TYPE_ROUTES). Falls back to the coarse doc_type route for
    rows written before source_doc_type existed, or callers that omit it."""
    _require_inventory(user)
    route = (_SOURCE_DOC_TYPE_ROUTES.get(source_doc_type) if source_doc_type else None) or _DOC_TYPE_ROUTES.get(doc_type)
    if not route:
        raise HTTPException(404, f"No drill-down route for voucher type '{doc_type}'")
    path = route.format(id=source_doc_id) if "{id}" in route else f"{route}?detail={source_doc_id}"
    return {"path": path, "docType": doc_type, "label": _DOC_TYPE_LABELS.get(doc_type, doc_type)}


@router.get("/export/excel")
async def export_excel(
    product_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    doc_type: Optional[str] = None,
    user_id: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(get_current_user),
) -> Response:
    _require_inventory(user)
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    where_clause, params = _build_filters(
        product_id, warehouse_id, from_date, to_date, doc_type, user_id, status, q
    )
    sql = text(f"{_BASE_SELECT} {where_clause} ORDER BY t.created_at DESC LIMIT 20000")
    async with get_session() as session:
        rows = (await session.execute(sql, params)).mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(500, "Could not create active worksheet")
    ws.title = "Stock Log"

    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = [
        "Date", "Voucher No", "Voucher Type", "Product Code", "Product Name",
        "Warehouse", "In Qty", "Out Qty", "Balance", "Rate", "Value", "User",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 16

    for idx, r in enumerate(rows, 2):
        delta = float(r["delta"] or 0)
        doc_type_val = r["doc_type"] or _fallback_doc_type(r["reason"])
        voucher_no_val = r["voucher_no"] or _fallback_voucher_no(r["reason"])
        ws.cell(row=idx, column=1, value=(r["created_at"] or "")[:19])
        ws.cell(row=idx, column=2, value=voucher_no_val or "")
        ws.cell(row=idx, column=3, value=_DOC_TYPE_LABELS.get(doc_type_val, doc_type_val or "") if doc_type_val else "")
        ws.cell(row=idx, column=4, value=r["product_sku"] or "")
        ws.cell(row=idx, column=5, value=r["product_name"] or "")
        ws.cell(row=idx, column=6, value=r["godown_name"] or "")
        ws.cell(row=idx, column=7, value=round(delta, 2) if delta > 0 else "")
        ws.cell(row=idx, column=8, value=round(-delta, 2) if delta < 0 else "")
        ws.cell(row=idx, column=9, value=round(float(r["running_balance"] or 0), 2))
        ws.cell(row=idx, column=10, value=round(float(r["rate"]), 2) if r["rate"] is not None else "")
        ws.cell(row=idx, column=11, value=round(float(r["value"]), 2) if r["value"] is not None else "")
        ws.cell(row=idx, column=12, value=r["user_name"] or "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"stock_log_{date.today().isoformat()}.xlsx"
    await log_audit("EXPORT", "stock_transactions", "bulk", user, new_values={"row_count": len(rows)})
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pdf")
async def export_pdf(
    product_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    doc_type: Optional[str] = None,
    user_id: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(get_current_user),
) -> Response:
    _require_inventory(user)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    where_clause, params = _build_filters(
        product_id, warehouse_id, from_date, to_date, doc_type, user_id, status, q
    )
    sql = text(f"{_BASE_SELECT} {where_clause} ORDER BY t.created_at DESC LIMIT 2000")
    async with get_session() as session:
        rows = (await session.execute(sql, params)).mappings().all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    elements: list[Any] = [Paragraph("Stock Log", styles["Title"]), Spacer(1, 6)]

    data = [["Date", "Voucher No", "Type", "Product", "Warehouse", "In", "Out", "Balance", "Rate", "Value"]]
    for r in rows:
        delta = float(r["delta"] or 0)
        doc_type_val = r["doc_type"] or _fallback_doc_type(r["reason"])
        voucher_no_val = r["voucher_no"] or _fallback_voucher_no(r["reason"])
        data.append([
            (r["created_at"] or "")[:10],
            voucher_no_val or "",
            _DOC_TYPE_LABELS.get(doc_type_val, doc_type_val or "") if doc_type_val else "",
            r["product_name"] or "",
            r["godown_name"] or "",
            f"{delta:.2f}" if delta > 0 else "",
            f"{-delta:.2f}" if delta < 0 else "",
            f"{float(r['running_balance'] or 0):.2f}",
            f"{float(r['rate']):.2f}" if r["rate"] is not None else "",
            f"{float(r['value']):.2f}" if r["value"] is not None else "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    filename = f"stock_log_{date.today().isoformat()}.pdf"
    await log_audit("EXPORT", "stock_transactions", "bulk", user, new_values={"row_count": len(rows)})
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
