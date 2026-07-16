"""Product Category master — CRUD, validation, delete-guard, import/export.

Pure unit tests against a small in-memory async Mongo fake that supports the
query operators the router uses ($ne, $or, $regex with ^..$ and $options=i).
Exercises the real handlers in routers.categories and routers.inventory.
"""
import asyncio
import io
import re

import pytest
from fastapi import HTTPException

import core.db
import core.utils as utils


# ───────────────────────── in-memory async Mongo fake ─────────────────────────

_MISSING = object()


def _match_value(actual, expected):
    """actual is the field value, or _MISSING when the key is absent."""
    if isinstance(expected, dict):
        for op, val in expected.items():
            if op == "$exists":
                present = actual is not _MISSING
                if present != bool(val):
                    return False
            elif op == "$ne":
                if (actual if actual is not _MISSING else None) == val:
                    return False
            elif op == "$regex":
                opts = expected.get("$options", "")
                flags = re.IGNORECASE if "i" in opts else 0
                a = None if actual is _MISSING else actual
                if a is None or re.search(val, str(a), flags) is None:
                    return False
            elif op == "$options":
                continue
            elif op == "$in":
                if (actual if actual is not _MISSING else None) not in val:
                    return False
            elif op == "$nin":
                if (actual if actual is not _MISSING else None) in val:
                    return False
            else:
                return actual == expected
        return True
    a = None if actual is _MISSING else actual
    return a == expected


def _doc_matches(d, q):
    for k, v in q.items():
        if k == "$or":
            if not any(_doc_matches(d, sub) for sub in v):
                return False
            continue
        if not _match_value(d.get(k, _MISSING), v):
            return False
    return True


class _Cursor:
    def __init__(self, docs):
        self._docs = [dict(x) for x in docs]

    def sort(self, spec, direction=None):
        if isinstance(spec, str):
            spec = [(spec, direction or 1)]
        for field, dirn in reversed(spec):
            self._docs.sort(key=lambda d: (d.get(field) is None, d.get(field)), reverse=(dirn == -1))
        return self

    def skip(self, n):
        self._docs = self._docs[n:]
        return self

    def limit(self, n):
        self._docs = self._docs[:n]
        return self

    async def to_list(self, _n=None):
        return [dict(x) for x in self._docs]

    def __aiter__(self):
        self._it = iter(self._docs)
        return self

    async def __anext__(self):
        try:
            return dict(next(self._it))
        except StopIteration:
            raise StopAsyncIteration


class _Collection:
    def __init__(self):
        self.docs = []

    async def insert_one(self, doc, session=None):
        doc = dict(doc)
        doc.setdefault("_id", doc.get("id"))
        self.docs.append(doc)
        return type("R", (), {"inserted_id": doc.get("id")})()

    async def insert_many(self, docs, session=None):
        for d in docs:
            await self.insert_one(d)
        return type("R", (), {"inserted_ids": [d.get("id") for d in docs]})()

    async def find_one(self, q, projection=None, session=None):
        for d in self.docs:
            if _doc_matches(d, q):
                out = dict(d)
                out.pop("_id", None)
                return out
        return None

    def find(self, q=None, projection=None):
        q = q or {}
        return _Cursor([d for d in self.docs if _doc_matches(d, q)])

    async def count_documents(self, q):
        return len([d for d in self.docs if _doc_matches(d, q)])

    async def update_one(self, q, update, session=None, upsert=False):
        for d in self.docs:
            if _doc_matches(d, q):
                for k, v in update.get("$set", {}).items():
                    d[k] = v
                return type("R", (), {"modified_count": 1})()
        return type("R", (), {"modified_count": 0})()

    async def update_many(self, q, update, session=None):
        n = 0
        for d in self.docs:
            if _doc_matches(d, q):
                for k, v in update.get("$set", {}).items():
                    d[k] = v
                n += 1
        return type("R", (), {"modified_count": n})()

    async def delete_one(self, q, session=None):
        for i, d in enumerate(self.docs):
            if _doc_matches(d, q):
                del self.docs[i]
                return type("R", (), {"deleted_count": 1})()
        return type("R", (), {"deleted_count": 0})()

    async def create_index(self, *a, **k):
        return None


class _DB:
    def __init__(self):
        self._cols = {}

    def __getitem__(self, name):
        return self._cols.setdefault(name, _Collection())

    def __getattr__(self, name):
        if name.startswith("_"):
            raise AttributeError(name)
        return self[name]


async def mock_crud_create(collection: str, data: dict, user: dict | None = None) -> dict:
    if "id" not in data or not data["id"]:
        data["id"] = utils.new_id()
    doc = dict(data)
    # Pack extra fields if table has extra (simulating schema)
    if collection == "product_categories":
        # Simulate postgres columns
        pass
    doc.setdefault("created_at", utils.now_iso())
    doc.setdefault("updated_at", utils.now_iso())
    from core.db import db as mock_db
    await mock_db[collection].insert_one(doc)
    return doc

async def mock_crud_get(collection: str, doc_id: str, label: str = "Record") -> dict:
    from core.db import db as mock_db
    doc = await mock_db[collection].find_one({"id": doc_id})
    if doc is None:
        raise HTTPException(status_code=404, detail=f"{label} not found")
    return doc

async def mock_crud_update(
    collection: str, doc_id: str, updates: dict, user: dict | None = None, label: str = "Record"
) -> dict:
    from core.db import db as mock_db
    updates = dict(updates)
    updates["updated_at"] = utils.now_iso()
    await mock_db[collection].update_one({"id": doc_id}, {"$set": updates})
    return await mock_crud_get(collection, doc_id, label=label)

async def mock_product_counts_by_category(rows: list[dict]) -> dict[str, int]:
    if not rows:
        return {}
    counts = {r["id"]: 0 for r in rows if r.get("id")}
    from core.db import db as mock_db
    products = await mock_db.products.find().to_list(1000)
    for p in products:
        cid = p.get("category_id")
        if cid in counts:
            counts[cid] += 1
    return counts


def _setup():
    import core.cache
    import core.db
    core.cache.clear()
    from typing import Any
    db: Any = _DB()
    core.db.db = db
    utils.db = db
    import routers.categories as cat
    cat.db = db
    cat.crud_create = mock_crud_create
    cat.crud_get = mock_crud_get
    cat.crud_update = mock_crud_update
    cat._product_counts_by_category = mock_product_counts_by_category
    import routers.inventory as inv
    inv.db = db
    inv.crud_create = mock_crud_create
    inv.crud_get = mock_crud_get
    inv.crud_update = mock_crud_update
    return db, cat, inv


USER = {"id": "u1", "name": "T", "role": "admin"}


def _make_payload(cat, **kw):
    from core.models import ProductCategory
    from typing import Any
    defaults: dict[str, Any] = dict(name="X", code=None, description=None, parent_id=None,
                                    status="Active", display_order=0)
    defaults.update(kw)
    return ProductCategory(**defaults)


# ───────────────────────── CRUD + validation ─────────────────────────

def test_create_autogenerates_code_and_lists():
    db, cat, _ = _setup()
    created = asyncio.run(cat.create_category(_make_payload(cat, name="Cuplock"), user=USER))
    assert created["name"] == "Cuplock"
    assert created["code"]  # auto-generated
    assert created["status"] == "Active"
    rows = asyncio.run(cat.list_categories(_=USER))
    assert len(rows) == 1
    assert rows[0]["product_count"] == 0


def test_duplicate_name_rejected_case_insensitive():
    db, cat, _ = _setup()
    asyncio.run(cat.create_category(_make_payload(cat, name="Frame"), user=USER))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(cat.create_category(_make_payload(cat, name="frame"), user=USER))
    assert exc.value.status_code == 400
    assert "already exists" in exc.value.detail


def test_codes_are_unique():
    db, cat, _ = _setup()
    a = asyncio.run(cat.create_category(_make_payload(cat, name="Base Jack"), user=USER))
    b = asyncio.run(cat.create_category(_make_payload(cat, name="Box Jacket"), user=USER))
    # Both slug to "BJ" → second must be disambiguated.
    assert a["code"] != b["code"]


def test_parent_must_exist():
    db, cat, _ = _setup()
    with pytest.raises(HTTPException) as exc:
        asyncio.run(cat.create_category(_make_payload(cat, name="Child", parent_id="nope"), user=USER))
    assert exc.value.status_code == 404


def test_update_renames_linked_products():
    db, cat, inv = _setup()
    c = asyncio.run(cat.create_category(_make_payload(cat, name="Planks"), user=USER))
    asyncio.run(db.products.insert_one({"id": "p1", "name": "Steel Plank", "sku": "SP1",
                                        "category": "Planks", "category_id": c["id"]}))
    from core.models import ProductCategoryUpdate
    asyncio.run(cat.update_category(c["id"], ProductCategoryUpdate(name="Steel Planks"), user=USER))
    prod = asyncio.run(db.products.find_one({"id": "p1"}))
    assert prod["category"] == "Steel Planks"


# ───────────────────────── delete guard (soft delete) ─────────────────────────

def test_delete_soft_deletes_when_unused():
    db, cat, _ = _setup()
    c = asyncio.run(cat.create_category(_make_payload(cat, name="Couplers"), user=USER))
    res = asyncio.run(cat.delete_category(c["id"], user=USER))
    assert res["soft_deleted"] is True
    # Soft-deleted rows drop out of the list.
    assert asyncio.run(cat.list_categories(_=USER)) == []
    raw = asyncio.run(db.product_categories.find_one({"id": c["id"]}))
    assert raw["is_deleted"] is True and raw["status"] == "Inactive"


def test_delete_blocked_when_products_linked_by_id():
    db, cat, _ = _setup()
    c = asyncio.run(cat.create_category(_make_payload(cat, name="Ringlock"), user=USER))
    asyncio.run(db.products.insert_one({"id": "p1", "name": "RL", "sku": "RL1",
                                        "category": "Ringlock", "category_id": c["id"]}))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(cat.delete_category(c["id"], user=USER))
    assert exc.value.status_code == 400
    assert "linked" in exc.value.detail


def test_delete_blocked_when_products_linked_by_legacy_name():
    db, cat, _ = _setup()
    c = asyncio.run(cat.create_category(_make_payload(cat, name="Accessories"), user=USER))
    # Legacy product with only the name set (no category_id).
    asyncio.run(db.products.insert_one({"id": "p1", "name": "Clamp", "sku": "C1",
                                        "category": "Accessories"}))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(cat.delete_category(c["id"], user=USER))
    assert "linked" in exc.value.detail


def test_delete_blocked_when_has_children():
    db, cat, _ = _setup()
    parent = asyncio.run(cat.create_category(_make_payload(cat, name="Parent"), user=USER))
    asyncio.run(cat.create_category(_make_payload(cat, name="Kid", parent_id=parent["id"]), user=USER))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(cat.delete_category(parent["id"], user=USER))
    assert "sub-categories" in exc.value.detail


# ───────────────────────── product ↔ category linkage ─────────────────────────

def test_product_create_links_category_by_id():
    db, cat, inv = _setup()
    c = asyncio.run(cat.create_category(_make_payload(cat, name="Cuplock"), user=USER))
    from core.models import Product
    p = Product(name="Vertical", sku="V1", category="ignored", category_id=c["id"])
    created = asyncio.run(inv.create_product(p, _=USER))
    assert created["category_id"] == c["id"]
    assert created["category"] == "Cuplock"  # canonical name copied from master


def test_product_create_links_category_by_name():
    db, cat, inv = _setup()
    c = asyncio.run(cat.create_category(_make_payload(cat, name="Frame"), user=USER))
    from core.models import Product
    p = Product(name="Frame Panel", sku="FP1", category="frame")  # case-insensitive
    created = asyncio.run(inv.create_product(p, _=USER))
    assert created["category_id"] == c["id"]


def test_product_create_unknown_category_id_rejected():
    db, cat, inv = _setup()
    from core.models import Product
    p = Product(name="X", sku="X1", category="X", category_id="missing")
    with pytest.raises(HTTPException) as exc:
        asyncio.run(inv.create_product(p, _=USER))
    assert exc.value.status_code == 400


# ───────────────────────── import / export ─────────────────────────

def _UploadFile(filename: str, data: bytes):
    from fastapi import UploadFile
    return UploadFile(filename=filename, file=io.BytesIO(data))


def test_import_csv_creates_categories():
    db, cat, _ = _setup()
    csv_bytes = (
        "Name,Code,Description,Status,Display Order\n"
        "Cuplock,,Main system,Active,1\n"
        "Ringlock,RNG,,Active,2\n"
    ).encode("utf-8")
    res = asyncio.run(cat.import_categories(
        file=_UploadFile("cats.csv", csv_bytes), commit=True, user=USER))
    assert res["valid_rows"] == 2
    assert res["committed"] == 2
    rows = asyncio.run(cat.list_categories(_=USER))
    assert {r["name"] for r in rows} == {"Cuplock", "Ringlock"}


def test_import_reports_duplicates_and_missing_name():
    db, cat, _ = _setup()
    asyncio.run(cat.create_category(_make_payload(cat, name="Cuplock"), user=USER))
    csv_bytes = (
        "Name,Code\n"
        "Cuplock,CUP\n"   # duplicate of existing
        ",BLANK\n"        # row present but name blank
        "Frame,FRM\n"     # ok
    ).encode("utf-8")
    res = asyncio.run(cat.import_categories(
        file=_UploadFile("c.csv", csv_bytes), commit=True, user=USER))
    assert res["valid_rows"] == 1
    assert res["error_rows"] == 2
    assert res["committed"] == 1


def test_import_dry_run_does_not_commit():
    db, cat, _ = _setup()
    csv_bytes = b"Name\nCuplock\n"
    res = asyncio.run(cat.import_categories(
        file=_UploadFile("c.csv", csv_bytes), commit=False, user=USER))
    assert res["valid_rows"] == 1
    assert res["committed"] == 0
    assert res["dry_run"] is True
    assert asyncio.run(cat.list_categories(_=USER)) == []


def test_export_returns_xlsx():
    db, cat, _ = _setup()
    asyncio.run(cat.create_category(_make_payload(cat, name="Cuplock"), user=USER))
    asyncio.run(cat.create_category(_make_payload(cat, name="Frame"), user=USER))
    resp = asyncio.run(cat.export_categories(_=USER))
    assert resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.body))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Name"
    names = {str(r[0]) for r in rows[1:]}
    assert names == {"Cuplock", "Frame"}


def test_migration_backfills_category_id():
    """Migration 017 links legacy products (category name, no id) to categories,
    creating any missing category, and is idempotent / skips already-linked rows."""
    db, cat, _ = _setup()
    import core.utils as u
    u.db = db  # the migration uses core.utils.new_id/now_iso (pure) + db indirectly

    # One pre-existing category, plus legacy products with only a name.
    existing = asyncio.run(cat.create_category(_make_payload(cat, name="Cuplock"), user=USER))
    asyncio.run(db.products.insert_one({"id": "p1", "name": "A", "sku": "A1", "category": "Cuplock"}))
    asyncio.run(db.products.insert_one({"id": "p2", "name": "B", "sku": "B1", "category": "Brand New"}))
    # Already linked — must be left untouched.
    asyncio.run(db.products.insert_one({"id": "p3", "name": "C", "sku": "C1",
                                        "category": "Cuplock", "category_id": "preset"}))

    from migrations.migration_017_product_category_link import run as run_017
    asyncio.run(run_017(db))

    p1 = asyncio.run(db.products.find_one({"id": "p1"}))
    p2 = asyncio.run(db.products.find_one({"id": "p2"}))
    p3 = asyncio.run(db.products.find_one({"id": "p3"}))
    assert p1["category_id"] == existing["id"]      # matched existing category
    assert p3["category_id"] == "preset"            # already-linked left alone
    # A new category was created for the previously-unknown name.
    new_cat = asyncio.run(db.product_categories.find_one({"name": "Brand New"}))
    assert new_cat is not None
    assert p2["category_id"] == new_cat["id"]

    # Idempotent: a second run changes nothing.
    asyncio.run(run_017(db))
    assert asyncio.run(db.products.find_one({"id": "p2"}))["category_id"] == new_cat["id"]


def test_import_xlsx_roundtrip():
    db, cat, _ = _setup()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Name", "Status", "Display Order"])
    ws.append(["Steel Prop", "Active", 5])
    buf = io.BytesIO()
    wb.save(buf)
    res = asyncio.run(cat.import_categories(
        file=_UploadFile("c.xlsx", buf.getvalue()), commit=True, user=USER))
    assert res["committed"] == 1
    rows = asyncio.run(cat.list_categories(_=USER))
    assert rows[0]["name"] == "Steel Prop"
    assert rows[0]["display_order"] == 5
